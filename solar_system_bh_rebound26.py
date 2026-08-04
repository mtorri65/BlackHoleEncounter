#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Solar System + Black Hole flyby with asteroid-belt tracers.

Features:
- Robust hyperbolic Kepler solver for BH.
- Hybrid symplectic IAS15 via Mercurius if available.
- Parameter sweeps over:
    * bh_rp_au / bh_rp_au_range
    * bh_vinf_kms / bh_vinf_kms_range
    * bh_inc_deg / bh_inc_deg_range
    * bh_tperi_offset_days / bh_tperi_offset_days_range
    * bh_Omega_deg / bh_Omega_deg_range
    * bh_omega_deg / bh_omega_deg_range
- Asteroid belt perturbations and hazard post-processing.
- planets_run_deltas.csv for 8 planets + Moon (no BH).
- Heliocentric plots:
    * Full system (existing)
    * Inner 3 AU (new)
- Inner-system animations (Mercury–Mars + Moon):
    * 2D x–y animation (MP4 if ffmpeg available; otherwise skipped gracefully)
    * 3D animation (same behavior)
- Optional SimulationArchive (for replay/post-viewing):
    * archive_enable
    * archive_interval_days
    * archive_filename

Resuming:
- Use --resume-dir simulations/YYYYMMDD_HHMMSS
  to resume a previous sweep. The script:
    * Reconstructs the expected parameter grid from input.yaml.
    * Finds the last matching subfolder in that root directory.
    * Deletes that last folder.
    * Re-runs that case and all remaining ones.
"""

import math
import os
import sys
import time
import traceback
import yaml
import argparse
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")  # headless / non-interactive backend
import matplotlib.pyplot as plt
from matplotlib import animation
from mpl_toolkits.mplot3d import Axes3D  # noqa: F401 (needed for 3D projection)
from pathlib import Path
from datetime import datetime, timezone
import shutil
from itertools import product
from datetime import timedelta
from concurrent.futures import ProcessPoolExecutor, as_completed
from openpyxl.utils import get_column_letter

try:
    import psutil
except ImportError:
    psutil = None


import rebound
from skyfield.api import load
from skyfield.constants import AU_KM

# ===========================================================
# Constants and planetary data
# ===========================================================
G_AU3_Msun_day2 = 0.0002959122082855911  # k^2 in AU^3/(Msun*day^2)
MSUN = 1.0
AU_M = AU_KM * 1000.0
AU_PER_DAY2_TO_M_S2 = AU_M / (86400.0 ** 2)

PLANET_MASSES_MSUN = {
    "mercury": 1.651e-7, "venus": 2.447e-6, "earth": 3.003e-6, "mars": 3.227e-7,
    "jupiter": 9.5479e-4, "saturn": 2.8585e-4, "uranus": 4.3662e-5, "neptune": 5.1513e-5,
}

# Skyfield targets (de440s: inner planets as centers, outers as barycenters)
SKYFIELD_KEYS = {
    "sun": "sun",
    "mercury": "mercury",
    "venus": "venus",
    "earth": "earth",
    "mars": "mars barycenter",
    "jupiter": "jupiter barycenter",
    "saturn": "saturn barycenter",
    "uranus": "uranus barycenter",
    "neptune": "neptune barycenter",
}

# Order we’ll log/plot in (indices in REBOUND: 0..10)
BODY_NAMES = [
    "Sun",
    "Mercury",
    "Venus",
    "Earth",
    "Mars",
    "Jupiter",
    "Saturn",
    "Uranus",
    "Neptune",
    "Moon",
    "BH",
]
PLANET_NAMES_FOR_DELTAS = ["Mercury","Venus","Earth","Mars","Jupiter","Saturn","Uranus","Neptune","Moon"]
INNER_BODY_NAMES = ["Mercury", "Venus", "Earth", "Mars", "Moon"]

# ===========================================================
# Particle lookup helpers (hash-first, index fallback)
# ===========================================================
def get_particle(sim, name: str, fallback_index: int = None):
    """Return particle by hash/name if present, otherwise by fallback index."""
    try:
        return sim.particles[name]
    except Exception:
        if fallback_index is None:
            raise
        return sim.particles[fallback_index]


# ===========================================================
# Small helpers
# ===========================================================
def _bh_params_box_text(params: dict) -> str:
    # Use the exact labels you requested
    return (
        f"Periapsis distance [AU]: {params['bh_rp_au']}\n"
        f"V_infinity [km/s]: {params['bh_vinf_kms']}\n"
        f"Inclination i [deg]: {params.get('bh_inc_deg', 0.0)}\n"
        f"Perihelion offset [days]: {params['bh_tperi_offset_days']}\n"
        f"Longitude of ascending node Ω [deg]: {params.get('bh_Omega_deg', 0.0)}\n"
        f"Argument of periapsis ω [deg]: {params.get('bh_omega_deg', 0.0)}"
    )

def _add_params_box(ax, params: dict, loc=(0.02, 0.98)):
    """
    Draw a small parameter box inside the axes.
    loc is in axes-fraction coordinates (0..1). Default is top-left.
    """
    txt = _bh_params_box_text(params)
    ax.text(
        loc[0], loc[1], txt,
        transform=ax.transAxes,
        va="top", ha="left",
        fontsize=8,
        bbox=dict(boxstyle="round,pad=0.3", facecolor="white", alpha=0.85, edgecolor="0.6")
    )

def parse_epoch_utc(s: str) -> datetime:
    s = s.strip()
    if s.endswith("Z"):
        return datetime.fromisoformat(s[:-1]).replace(tzinfo=timezone.utc)
    dt = datetime.fromisoformat(s)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc)


def safe_num_for_filename(x: float) -> str:
    """Convert numeric value to a filename-safe token: 3.5 -> 3p5, -1.0 -> m1p0"""
    s = f"{x:g}"
    return s.replace("-", "m").replace(".", "p")

def _fmt_full_precision(x):
    """Return a float formatted with full precision for CSV logging."""
    try:
        return f"{float(x):.15e}"
    except Exception:
        return ""    

def _parse_range(val):
    """
    Accepts:
      - string: "min,max,step"
      - list/tuple: [min, max, step]
    Returns a sorted numpy array including the upper bound (within tolerance).
    """
    if isinstance(val, str):
        parts = [float(p.strip()) for p in val.split(",")]
    elif isinstance(val, (list, tuple)) and len(val) == 3:
        parts = [float(val[0]), float(val[1]), float(val[2])]
    else:
        raise ValueError("Range must be 'min,max,step' string or [min,max,step] list/tuple.")
    lo, hi, step = parts
    if step <= 0:
        raise ValueError("Range step must be > 0.")
    n = int(round((hi - lo) / step)) + 1
    arr = lo + step * np.arange(max(n, 1))
    arr = arr[(arr >= lo - 1e-12) & (arr <= hi + 1e-12)]
    return np.round(arr, 12)


def _prefix_path(base_dir: Path, name: str) -> Path:
    """Return a path inside base_dir, prefixed with base_dir.name + '__'."""
    return base_dir / f"{base_dir.name}__{name}"


# >>> RA/Dec helper (equatorial frame) <<<
def vector_to_radec(vec_eq: np.ndarray):
    """
    Convert a 3D ICRF/equatorial vector to (RA_hours, Dec_deg).

    Parameters
    ----------
    vec_eq : np.ndarray
        [x, y, z] in ICRF/equatorial frame.

    Returns
    -------
    ra_hours : float
        Right Ascension in hours, 0..24.
    dec_deg : float
        Declination in degrees, -90..+90.
    """
    x, y, z = vec_eq
    r = float(np.linalg.norm(vec_eq))
    if r == 0.0:
        raise ValueError("Zero-length vector: cannot compute RA/Dec.")

    # unit vector
    x /= r
    y /= r
    z /= r

    dec_rad = math.asin(z)
    ra_rad = math.atan2(y, x)
    if ra_rad < 0.0:
        ra_rad += 2.0 * math.pi

    ra_hours = (math.degrees(ra_rad) / 15.0)  # 360 deg = 24h
    # guard against roundoff producing 24.000...
    if ra_hours >= 24.0:
        ra_hours -= 24.0

    dec_deg = math.degrees(dec_rad)
    return ra_hours, dec_deg

# ===========================================================
# Orbital-element helpers for belt snapshots (heliocentric)
# ===========================================================
def _elem_from_state_heliocentric(x, y, z, vx, vy, vz, mu):
    """
    Vectorized computation of a, e, i, q from heliocentric state.
    """
    x = np.array(x, dtype=float)
    y = np.array(y, dtype=float)
    z = np.array(z, dtype=float)
    vx = np.array(vx, dtype=float)
    vy = np.array(vy, dtype=float)
    vz = np.array(vz, dtype=float)

    r = np.sqrt(x * x + y * y + z * z)
    v2 = vx * vx + vy * vy + vz * vz

    hx = y * vz - z * vy
    hy = z * vx - x * vz
    hz = x * vy - y * vx
    h2 = hx * hx + hy * hy + hz * hz
    h = np.sqrt(h2)

    hz_over_h = np.clip(np.where(h > 0, hz / h, 1.0), -1.0, 1.0)
    inc = np.arccos(hz_over_h)

    inv_a = (2.0 / np.maximum(r, 1e-30)) - (v2 / np.maximum(mu, 1e-30))
    a = np.where(np.abs(inv_a) > 1e-30, 1.0 / inv_a, np.sign(inv_a) * 1e30)

    cx = vy * hz - vz * hy
    cy = vz * hx - vx * hz
    cz = vx * hy - vy * hx
    rx_hat = np.where(r > 0, x / r, 0.0)
    ry_hat = np.where(r > 0, y / r, 0.0)
    rz_hat = np.where(r > 0, z / r, 0.0)
    ex = cx / np.maximum(mu, 1e-30) - rx_hat
    ey = cy / np.maximum(mu, 1e-30) - ry_hat
    ez = cz / np.maximum(mu, 1e-30) - rz_hat
    e = np.sqrt(ex * ex + ey * ey + ez * ez)

    q_ell = a * (1.0 - e)
    q_hyp = np.where(
        (1.0 + e) > 1e-12,
        h2 / (np.maximum(mu, 1e-30) * (1.0 + e)),
        np.nan,
    )
    q = np.where(e < 1.0, q_ell, q_hyp)

    return a, e, inc, q


def snapshot_belt(sim, belt_start_index, belt_count, label, mu):
    """Return DataFrame of a/e/i/q (heliocentric) for belt tracers."""
    sun = sim.particles[0]
    xs = []
    ys = []
    zs = []
    vxs = []
    vys = []
    vzs = []
    for i in range(belt_start_index, belt_start_index + belt_count):
        p = sim.particles[i]
        xs.append(p.x - sun.x)
        ys.append(p.y - sun.y)
        zs.append(p.z - sun.z)
        vxs.append(p.vx - sun.vx)
        vys.append(p.vy - sun.vy)
        vzs.append(p.vz - sun.vz)
    x, y, z, vx, vy, vz = map(np.array, (xs, ys, zs, vxs, vys, vzs))
    a, e, inc, q = _elem_from_state_heliocentric(x, y, z, vx, vy, vz, mu)
    return pd.DataFrame(
        {
            f"a_{label}": a,
            f"e_{label}": e,
            f"i_{label}": inc,
            f"q_{label}": q,
        }
    )


def snapshot_planets_heliocentric(sim, mu, body_names=PLANET_NAMES_FOR_DELTAS):
    """
    Get heliocentric orbital elements (a,e,q,Q) for selected massive bodies
    (8 planets + Moon; BH excluded).

    Returns a DataFrame with columns:
      body, a, e, q, Q
    """    # Use hashes when available; index fallback for backward-compatibility
    sun = get_particle(sim, "Sun", fallback_index=0)
    rows = []
    for nm in body_names:
        p = get_particle(sim, nm, fallback_index=BODY_NAMES.index(nm))
        x = p.x - sun.x
        y = p.y - sun.y
        z = p.z - sun.z
        vx = p.vx - sun.vx
        vy = p.vy - sun.vy
        vz = p.vz - sun.vz
        a, e, inc, q = _elem_from_state_heliocentric(
            np.array([x]),
            np.array([y]),
            np.array([z]),
            np.array([vx]),
            np.array([vy]),
            np.array([vz]),
            mu,
        )
        a_val = float(a[0])
        e_val = float(e[0])
        q_val = float(q[0])
        Q_val = a_val * (1.0 + e_val)
        rows.append(
            {
                "body": nm,
                "a_AU": a_val,
                "e": e_val,
                "q_AU": q_val,
                "Q_AU": Q_val,
            }
        )
    return pd.DataFrame(rows)


# ===========================================================
# Build the REBOUND simulation
# ===========================================================
def build_sim(params):
    epoch_iso = params["epoch"]
    kernel = params["kernel"]
    bh_mass = params["bh_mass_msun"]
    bh_rp = params["bh_rp_au"]
    bh_vinf = params["bh_vinf_kms"]
    bh_toff = params["bh_tperi_offset_days"]
    bh_Omega = math.radians(params.get("bh_Omega_deg", 0.0))
    bh_inc = math.radians(params.get("bh_inc_deg", 0.0))
    bh_omega = math.radians(params.get("bh_omega_deg", 0.0))

    # Load ephemeris/time
    eph = load(f"{kernel}.bsp" if not kernel.lower().endswith(".bsp") else kernel)
    ts = load.timescale()
    t0 = ts.from_datetime(parse_epoch_utc(epoch_iso))

    # Sun barycentric state to build heliocentric states
    sun_bary = eph[SKYFIELD_KEYS["sun"]].at(t0)
    sun_r, sun_v = sun_bary.position.au, sun_bary.velocity.au_per_d

    def heliocentric_state(key):
        b = eph[key].at(t0)
        return b.position.au - sun_r, b.velocity.au_per_d - sun_v

    sim = rebound.Simulation()
    sim.G = G_AU3_Msun_day2

    # --- Hybrid integrator: WHFast default, IAS15 during close encounters ---
    dt_days = float(params.get("dt_days", 0.5))
    sim.dt = dt_days
    sim.integrator = "ias15"


    # Sun
    sim.add(m=MSUN, name="Sun")

    # Planets
    order = ["mercury", "venus", "earth", "mars", "jupiter", "saturn", "uranus", "neptune"]
    for nm in order:
        r, v = heliocentric_state(SKYFIELD_KEYS[nm])
        sim.add(
            m=PLANET_MASSES_MSUN[nm],
            x=r[0],
            y=r[1],
            z=r[2],
            vx=v[0],
            vy=v[1],
            vz=v[2],
            name=nm.capitalize(),  # Mercury, Venus, Earth, ...
        )

    # Moon (simple approx in Earth's orbital plane / near it)
    earth_idx = order.index("earth") + 1
    earth = get_particle(sim, "Earth", fallback_index=earth_idx)
    moon_r_au = 384400.0 / AU_KM
    sim.add(
        m=3.694e-8,
        x=earth.x + moon_r_au,
        y=earth.y,
        z=earth.z,
        vx=earth.vx,
        vy=earth.vy + 0.00059,
        vz=earth.vz,
        name="Moon"
    )

    # Black hole on hyperbolic approach relative to Sun
    mu_sun_bh = sim.G * (MSUN + bh_mass)
    vinf_au_per_day = bh_vinf / (AU_KM / 86400.0)
    a = -mu_sun_bh / (vinf_au_per_day ** 2)
    e = 1.0 + bh_rp / abs(a)
    n = math.sqrt(mu_sun_bh / abs(a) ** 3)
    M_epoch = -n * bh_toff

    # ---------- Robust hyperbolic Kepler solver ----------
    def solve_kepler_hyperbolic(M, e, tol=1e-13, max_iter=100):
        # Initial guess: use asinh(M/e) which is defined for all M
        F = math.asinh(M / max(e, 1e-16))
        for _ in range(max_iter):
            s = math.sinh(F)
            c = math.cosh(F)
            f = e * s - F - M
            fp = e * c - 1.0
            dF = -f / fp
            if abs(dF) > 1.0:
                dF = math.copysign(1.0, dF)
            F += dF
            if abs(dF) < tol:
                break
        return F

    F = solve_kepler_hyperbolic(M_epoch, e)
    coshF, sinhF = math.cosh(F), math.sinh(F)
    cosf = (e - coshF) / (e * coshF - 1.0)
    sinf = math.sqrt(max(e * e - 1.0, 0.0)) * sinhF / (e * coshF - 1.0)
    f = math.atan2(sinf, cosf)
    r = a * (e * coshF - 1.0)
    h = math.sqrt(mu_sun_bh * abs(a) * max(e * e - 1.0, 0.0))
    x_pqw = r * np.array([math.cos(f), math.sin(f), 0.0])
    v_pqw = (mu_sun_bh / h) * np.array([-math.sin(f), (e + math.cos(f)), 0.0])

    cO, sO = math.cos(bh_Omega), math.sin(bh_Omega)
    ci, si = math.cos(bh_inc), math.sin(bh_inc)
    co, so = math.cos(bh_omega), math.sin(bh_omega)
    R = np.array(
        [
            [cO * co - sO * so * ci, -cO * so - sO * co * ci, sO * si],
            [sO * co + cO * so * ci, -sO * so + cO * co * ci, -cO * si],
            [so * si, co * si, ci],
        ]
    )
    r_bh = R @ x_pqw
    v_bh = -(R @ v_pqw)  # inbound branch

    sim.add(
        m=bh_mass,
        x=r_bh[0],
        y=r_bh[1],
        z=r_bh[2],
        vx=v_bh[0],
        vy=v_bh[1],
        vz=v_bh[2],
        name="BH",
    )

    # Asteroid belt tracers (massless)
    n_belt = int(params.get("n_belt", 4000))
    rng = np.random.default_rng(42)
    belt_start = sim.N
    a_b = rng.uniform(2.0, 3.5, n_belt)
    e_b = rng.uniform(0.0, 0.15, n_belt)
    i_b = np.radians(rng.uniform(0.0, 10.0, n_belt))
    M_b = rng.uniform(0.0, 2 * np.pi, n_belt)

    for k in range(n_belt):
        a_k, e_k, i_k, M_k = a_b[k], e_b[k], i_b[k], M_b[k]
        r_k = a_k * (1.0 - e_k * e_k) / (1.0 + e_k * math.cos(M_k))
        x = r_k * math.cos(M_k)
        y = r_k * math.sin(M_k)
        z = r_k * math.sin(i_k)
        v_circ = math.sqrt(G_AU3_Msun_day2 * MSUN * (2.0 / r_k - 1.0 / a_k))
        vx = -v_circ * math.sin(M_k)
        vy = v_circ * math.cos(M_k)
        vz = 0.0
        sim.add(m=0.0, x=x, y=y, z=z, vx=vx, vy=vy, vz=vz)

    belt = {"start": belt_start, "count": n_belt}

    # Sanity check: ensure key hashes are present (helps avoid anonymous archives)
    for _nm in ["Sun","Mercury","Venus","Earth","Mars","Jupiter","Saturn","Uranus","Neptune","Moon","BH"]:
        try:
            _ = sim.particles[_nm]
        except Exception as e:
            raise RuntimeError(f"Missing or non-unique particle hash: '{_nm}'.") from e

    sim.move_to_com()
    return sim, belt, mu_sun_bh

# ===========================================================
# Snapshots, logging, Excel, plots, animations
# ===========================================================
def write_snapshot(sim, output_dir: Path, t_days: float, names=None, write_npz=True):
    if not write_npz:
        return
    arr = np.zeros((sim.N, 6))
    for i, p in enumerate(sim.particles):
        arr[i] = [p.x, p.y, p.z, p.vx, p.vy, p.vz]
    fname = _prefix_path(output_dir, f"snapshot_{int(t_days):06d}.npz")
    np.savez_compressed(fname, arr=arr, names=np.array(names or [], dtype=object))


def init_logs(body_names):
    return {
        nm: {
            "t": [],
            "x": [],
            "y": [],
            "z": [],
            "vx": [],
            "vy": [],
            "vz": [],
            "r_helio_AU": [],
            "r_helio_AU_str": [],
            "a_tidal_AU_per_d2": [],
            "a_tidal_AU_per_d2_str": [],
            "a_tidal_m_s2": [],
            "a_tidal_m_s2_str": [],
            "disp_helio_AU": [],
            "disp_helio_AU_str": [],
            "disp_helio_m": [],
            "disp_helio_m_str": [],
        }
        for nm in body_names
    }


def record_logs(sim, logs, body_names, t, sim_baseline=None):
    """
    Record states and derived quantities for each logged body.

    Added in v25:
      - heliocentric distance r_helio_AU
      - BH tidal acceleration scalar estimate
      - true heliocentric displacement magnitude relative to an optional baseline sim
    """
    sun = get_particle(sim, "Sun", fallback_index=0)
    bh = get_particle(sim, "BH", fallback_index=BODY_NAMES.index("BH"))
    dxR, dyR, dzR = bh.x - sun.x, bh.y - sun.y, bh.z - sun.z
    R = math.sqrt(dxR * dxR + dyR * dyR + dzR * dzR)

    if R > 0.0:
        pref = (2.0 * sim.G * float(bh.m)) / (R ** 3)  # 1/day^2
    else:
        pref = float("nan")

    if sim_baseline is not None:
        sun0 = get_particle(sim_baseline, "Sun", fallback_index=0)
    else:
        sun0 = None

    for i, nm in enumerate(body_names):
        p = get_particle(sim, nm, fallback_index=i)
        logs[nm]["t"].append(t)
        logs[nm]["x"].append(p.x)
        logs[nm]["y"].append(p.y)
        logs[nm]["z"].append(p.z)
        logs[nm]["vx"].append(p.vx)
        logs[nm]["vy"].append(p.vy)
        logs[nm]["vz"].append(p.vz)

        dx = p.x - sun.x
        dy = p.y - sun.y
        dz = p.z - sun.z
        r_helio = math.sqrt(dx * dx + dy * dy + dz * dz)
        logs[nm]["r_helio_AU"].append(r_helio)
        logs[nm]["r_helio_AU_str"].append(_fmt_full_precision(r_helio))

        if nm == "BH":
            a_tidal = float("nan")
        else:
            a_tidal = pref * r_helio if np.isfinite(pref) else float("nan")
        a_tidal_ms2 = a_tidal * AU_PER_DAY2_TO_M_S2 if np.isfinite(a_tidal) else float("nan")
        logs[nm]["a_tidal_AU_per_d2"].append(a_tidal)
        logs[nm]["a_tidal_AU_per_d2_str"].append(_fmt_full_precision(a_tidal))
        logs[nm]["a_tidal_m_s2"].append(a_tidal_ms2)
        logs[nm]["a_tidal_m_s2_str"].append(_fmt_full_precision(a_tidal_ms2))

        if sim_baseline is not None:
            p0 = get_particle(sim_baseline, nm, fallback_index=i)
            dx0 = p0.x - sun0.x
            dy0 = p0.y - sun0.y
            dz0 = p0.z - sun0.z
            disp_helio_au = math.sqrt((dx - dx0) ** 2 + (dy - dy0) ** 2 + (dz - dz0) ** 2)
            disp_helio_m = disp_helio_au * AU_M
        else:
            disp_helio_au = float("nan")
            disp_helio_m = float("nan")

        logs[nm]["disp_helio_AU"].append(disp_helio_au)
        logs[nm]["disp_helio_AU_str"].append(_fmt_full_precision(disp_helio_au))
        logs[nm]["disp_helio_m"].append(disp_helio_m)
        logs[nm]["disp_helio_m_str"].append(_fmt_full_precision(disp_helio_m))

def integrate_and_dump(
    sim,
    duration_days,
    dump_interval_days,
    output_interval_days,
    output_dir,
    body_names,
    logs,
    epoch_dt_utc,        # required (your RA/Dec timestamps)
    write_npz=True,      # <<< NEW
    sim_baseline=None,
):

    if dump_interval_days <= 0:
        dump_interval_days = duration_days
    steps = int(math.ceil(duration_days / dump_interval_days))
    last_logged_t = -1e99

    # >>> BH RA/Dec storage (ICRF geocentric) <<<
    bh_radec = {"t_days": [], "utc_iso": [], "RA_hours": [], "Dec_deg": []}

    bh_radec.update({
            "v_para_AU_per_d": [],
            "v_perp_AU_per_d": [],
            "v_para_km_s": [],
            "v_perp_km_s": [],
            "range_AU": [],   # optional but often useful
        })
    
    earth_index = BODY_NAMES.index("Earth")
    bh_index = BODY_NAMES.index("BH")  # fallback index only

    for step in range(steps + 1):
        t_target = min(step * dump_interval_days, duration_days)
        sim.integrate(t_target)
        if sim_baseline is not None:
            sim_baseline.integrate(t_target)

        # >>> NPZ snapshots optional <<<
        write_snapshot(sim, output_dir, t_target, names=BODY_NAMES, write_npz=write_npz)

        if (t_target - last_logged_t) >= max(1e-9, output_interval_days):
            record_logs(sim, logs, BODY_NAMES, t_target, sim_baseline=sim_baseline)
            last_logged_t = t_target

            # >>> log BH geocentric RA/Dec at same cadence (ICRF sky coords) <<<
            earth = get_particle(sim, "Earth", fallback_index=earth_index)
            bh = get_particle(sim, "BH", fallback_index=bh_index)

            los_vec = np.array(
                [bh.x - earth.x, bh.y - earth.y, bh.z - earth.z],
                dtype=float,
            )
            r = float(np.linalg.norm(los_vec))

            ra_hours, dec_deg = vector_to_radec(los_vec)

            if r > 0.0:
                los_hat = los_vec / r
                v_rel = np.array([bh.vx - earth.vx, bh.vy - earth.vy, bh.vz - earth.vz], dtype=float)

                v_para = float(np.dot(v_rel, los_hat))  # AU/day, signed (+ receding)
                v_perp_vec = v_rel - v_para * los_hat
                v_perp = float(np.linalg.norm(v_perp_vec))  # AU/day, magnitude

                # Convert AU/day -> km/s
                AU_PER_DAY_TO_KM_S = AU_KM / 86400.0
                v_para_kms = v_para * AU_PER_DAY_TO_KM_S
                v_perp_kms = v_perp * AU_PER_DAY_TO_KM_S
            else:
                v_para = v_perp = float("nan")
                v_para_kms = v_perp_kms = float("nan")

            # UTC timestamp = epoch + t_days
            utc_dt = epoch_dt_utc + timedelta(days=float(t_target))

            bh_radec["t_days"].append(t_target)
            bh_radec["utc_iso"].append(utc_dt.strftime("%Y-%m-%dT%H:%M:%SZ"))
            bh_radec["RA_hours"].append(ra_hours)
            bh_radec["Dec_deg"].append(dec_deg)

            bh_radec["v_para_AU_per_d"].append(v_para)
            bh_radec["v_perp_AU_per_d"].append(v_perp)
            bh_radec["v_para_km_s"].append(v_para_kms)
            bh_radec["v_perp_km_s"].append(v_perp_kms)
            bh_radec["range_AU"].append(r)            

        if step % 10 == 0 or step == steps:
            bh = get_particle(sim, "BH", fallback_index=bh_index)
            sun = get_particle(sim, "Sun", fallback_index=0)
            dx, dy, dz = bh.x - sun.x, bh.y - sun.y, bh.z - sun.z
            r_bh_sun = math.sqrt(dx * dx + dy * dy + dz * dz)
            print(f"[t={t_target:9.3f} d] r_BH-Sun={r_bh_sun:.3f} AU")

    return bh_radec


def write_excel(base_dir: Path, params: dict, logs: dict, body_names):
    rp_token = safe_num_for_filename(params["bh_rp_au"])
    m_token = safe_num_for_filename(params["bh_mass_msun"])
    xlsx_path = _prefix_path(base_dir, f"orbits__{rp_token}__{m_token}.xlsx")
    try:
        with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
            summary_params = [
                ("Epoch (UTC)", params["epoch"]),
                ("Kernel", params["kernel"]),
                ("BH mass (Msun)", params["bh_mass_msun"]),
                ("BH rp (AU)", params["bh_rp_au"]),
                ("BH vinf (km/s)", params["bh_vinf_kms"]),
                ("BH t_peri offset (days)", params["bh_tperi_offset_days"]),
                ("BH Omega (deg)", params.get("bh_Omega_deg", 0.0)),
                ("BH inc (deg)", params.get("bh_inc_deg", 0.0)),
                ("BH omega (deg)", params.get("bh_omega_deg", 0.0)),
                ("Duration (days)", params["duration_days"]),
                ("Output interval (days)", params["output_interval_days"]),
                ("Dump interval (days)", params["dump_interval_days"]),
                ("n_belt", int(params.get("n_belt", 4000))),
                ("compare_baseline_bh_mass_msun", params.get("compare_baseline_bh_mass_msun", "")),
            ]
            summary = pd.DataFrame({"Parameter": [k for k, _ in summary_params], "Value": [v for _, v in summary_params]})
            summary.to_excel(writer, index=False, sheet_name="Summary")

            numeric_cols_to_format = {
                "r_helio_AU",
                "a_tidal_AU_per_d2",
                "a_tidal_m_s2",
                "disp_helio_AU",
                "disp_helio_m",
            }
            sci_fmt = "0.0000000000000000E+00"

            for nm in body_names:
                df = pd.DataFrame({
                    "time_days": logs[nm]["t"],
                    "x_AU": logs[nm]["x"],
                    "y_AU": logs[nm]["y"],
                    "z_AU": logs[nm]["z"],
                    "vx_AU_per_d": logs[nm]["vx"],
                    "vy_AU_per_d": logs[nm]["vy"],
                    "vz_AU_per_d": logs[nm]["vz"],
                    "r_helio_AU": logs[nm].get("r_helio_AU", []),
                    "r_helio_AU_str": logs[nm].get("r_helio_AU_str", []),
                    "a_tidal_AU_per_d2": logs[nm].get("a_tidal_AU_per_d2", []),
                    "a_tidal_AU_per_d2_str": logs[nm].get("a_tidal_AU_per_d2_str", []),
                    "a_tidal_m_s2": logs[nm].get("a_tidal_m_s2", []),
                    "a_tidal_m_s2_str": logs[nm].get("a_tidal_m_s2_str", []),
                    "disp_helio_AU": logs[nm].get("disp_helio_AU", []),
                    "disp_helio_AU_str": logs[nm].get("disp_helio_AU_str", []),
                    "disp_helio_m": logs[nm].get("disp_helio_m", []),
                    "disp_helio_m_str": logs[nm].get("disp_helio_m_str", []),
                })
                df.to_excel(writer, index=False, sheet_name=nm)

                ws = writer.book[nm]
                header_to_col = {cell.value: cell.column for cell in ws[1]}
                for col_name in numeric_cols_to_format:
                    if col_name in header_to_col:
                        col_idx = header_to_col[col_name]
                        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                            row[0].number_format = sci_fmt
        print(f"Excel written: {xlsx_path}")
    except Exception as e:
        print(f"(Excel write skipped: {e})")


# >>> New: write BH RA/Dec Excel <<<
def write_bh_radec_excel(base_dir: Path, params: dict, bh_radec: dict):
    """
    Save BH geocentric RA/Dec vs time to an Excel file.

    Columns:
      time_days, RA_deg, Dec_deg
    """
    if not bh_radec["t_days"]:
        print("No BH RA/Dec records; skipping bh_radec Excel.")
        return

    rp_token = safe_num_for_filename(params["bh_rp_au"])
    m_token = safe_num_for_filename(params["bh_mass_msun"])
    xlsx_path = _prefix_path(base_dir, f"bh_radec__{rp_token}__{m_token}.xlsx")

    df = pd.DataFrame(
        {
            "time_days": bh_radec["t_days"],
            "utc_iso": bh_radec["utc_iso"],
            "RA_hours": bh_radec["RA_hours"],
            "Dec_deg": bh_radec["Dec_deg"],
            "range_AU": bh_radec.get("range_AU", []),
            "v_para_AU_per_d": bh_radec.get("v_para_AU_per_d", []),
            "v_perp_AU_per_d": bh_radec.get("v_perp_AU_per_d", []),
            "v_para_km_s": bh_radec.get("v_para_km_s", []),
            "v_perp_km_s": bh_radec.get("v_perp_km_s", []),
        }
    )
    try:
        with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="BH_RADEC")
        print(f"BH RA/Dec Excel written: {xlsx_path}")
    except Exception as e:
        print(f"(BH RA/Dec Excel write skipped: {e})")


def write_uncertainty_csv(base_dir: Path, params: dict, logs: dict, body_name: str = "Earth"):
    """
    Export one or more CSV files directly consumable by
    estimate_BH_parameters_uncertainties_fast.py.

    Output columns in each CSV:
      - time_days
      - Delta_tidal (m)   <- disp_helio_m for the selected body
      - BH_Sun_d (AU)     <- BH heliocentric distance r_helio_AU

    By default, if params['uncertainty_bodies'] is not provided, a single file
    is written for ``body_name``. If params['uncertainty_bodies'] is set to
    ``all`` / ``all_planets`` / ``planets``, one CSV is written for each of the
    eight major planets. A custom list may also be supplied.

    Files are written to SolarSystem/uncertainties/ by default, or to the
    directory provided via params['uncertainty_output_dir'].
    """
    if "BH" not in logs:
        print("[uncertainty] BH log not found; skipping CSV export.")
        return

    raw_bodies = params.get("uncertainty_bodies", None)
    default_planets = [
        "Mercury", "Venus", "Earth", "Mars",
        "Jupiter", "Saturn", "Uranus", "Neptune",
    ]

    if raw_bodies is None:
        body_names = [str(body_name)]
    elif isinstance(raw_bodies, str):
        token = raw_bodies.strip().lower()
        if token in {"all", "all_planets", "planets"}:
            body_names = default_planets
        elif "," in raw_bodies:
            body_names = [b.strip() for b in raw_bodies.split(",") if b.strip()]
        else:
            body_names = [raw_bodies.strip()]
    elif isinstance(raw_bodies, (list, tuple, set)):
        body_names = [str(b).strip() for b in raw_bodies if str(b).strip()]
    else:
        body_names = [str(body_name)]

    # De-duplicate while preserving order
    seen = set()
    body_names = [b for b in body_names if not (b in seen or seen.add(b))]

    bh = logs["BH"]
    bh_sun_d_au = np.asarray(bh.get("r_helio_AU", []), dtype=float)

    rp_token = safe_num_for_filename(params["bh_rp_au"])
    m_token = safe_num_for_filename(params["bh_mass_msun"])
    inc_token = safe_num_for_filename(params.get("bh_inc_deg", 0.0))
    Omega_token = safe_num_for_filename(params.get("bh_Omega_deg", 0.0))

    out_dir = Path(params.get("uncertainty_output_dir", "uncertainties"))
    if not out_dir.is_absolute():
        out_dir = Path(__file__).resolve().parent / out_dir
    out_dir.mkdir(parents=True, exist_ok=True)

    filename_template = params.get("uncertainty_filename_template", None)
    single_filename = params.get("uncertainty_filename", None)

    wrote_any = False
    for this_body in body_names:
        if this_body not in logs:
            print(f"[uncertainty] Body '{this_body}' not found in logs; skipping CSV export.")
            continue

        body = logs[this_body]
        time_days = np.asarray(body.get("t", []), dtype=float)
        delta_tidal_m = np.asarray(body.get("disp_helio_m", []), dtype=float)

        n = min(len(time_days), len(delta_tidal_m), len(bh_sun_d_au))
        if n == 0:
            print(f"[uncertainty] No logged samples available for {this_body}; skipping CSV export.")
            continue

        if not (len(time_days) == len(delta_tidal_m) == len(bh_sun_d_au)):
            print(
                "[uncertainty] Warning: log lengths differ "
                f"for {this_body} (time={len(time_days)}, delta={len(delta_tidal_m)}, bh={len(bh_sun_d_au)}); "
                f"truncating to {n} rows."
            )

        body_token = str(this_body).strip().replace(" ", "_")
        default_name = (
            f"uncertainty_input__{body_token}"
            f"__rp{rp_token}"
            f"__m{m_token}"
            f"__inc{inc_token}"
            f"__Om{Omega_token}.csv"
        )

        if filename_template:
            filename = str(filename_template).format(
                body=body_token,
                rp=rp_token,
                m=m_token,
                inc=inc_token,
                Om=Omega_token,
            )
        elif single_filename and len(body_names) == 1:
            filename = single_filename
        else:
            filename = default_name

        out_path = out_dir / filename

        df = pd.DataFrame(
            {
                "time_days": time_days[:n],
                "Delta_tidal (m)": delta_tidal_m[:n],
                "BH_Sun_d (AU)": bh_sun_d_au[:n],
            }
        )
        df.to_csv(out_path, index=False)
        wrote_any = True

        if np.all(~np.isfinite(df["Delta_tidal (m)"].to_numpy(dtype=float))):
            print(
                f"[uncertainty] CSV written for {this_body}, but Delta_tidal is all NaN. "
                "Set compare_baseline_bh_mass_msun to enable disp_helio_m output."
            )
        print(f"Uncertainty CSV written: {out_path}")

    if not wrote_any:
        print("[uncertainty] No uncertainty CSV files were written.")

def plot_xy_heliocentric_from_logs(base_dir: Path, params: dict, logs: dict, body_names):
    sun_x = np.array(logs["Sun"]["x"])
    sun_y = np.array(logs["Sun"]["y"])

    def rel_xy(nm):
        x = np.array(logs[nm]["x"]); y = np.array(logs[nm]["y"])
        n = min(len(x), len(sun_x))
        return x[:n] - sun_x[:n], y[:n] - sun_y[:n]

    fig, ax = plt.subplots(figsize=(8, 8))
    ax.set_aspect("equal")
    ax.grid(True, alpha=0.3)
    ax.set_xlabel("x [AU]")
    ax.set_ylabel("y [AU]")
    ax.set_title("Solar System + BH flyby (heliocentric x–y)")

    # Parameter box
    _add_params_box(ax, params, loc=(0.02, 0.98))

    for nm in ["Mercury","Venus","Earth","Mars","Jupiter","Saturn","Uranus","Neptune","Moon"]:
        if nm not in logs or len(logs[nm]["x"]) == 0:
            continue
        xr, yr = rel_xy(nm)
        ax.plot(xr, yr, lw=1.0, label=nm)
        if len(xr) > 0:
            ax.scatter(xr[-1], yr[-1], s=10)

    # BH
    if "BH" in logs and len(logs["BH"]["x"]) > 0:
        xr, yr = rel_xy("BH")
        ax.plot(xr, yr, lw=2.0, label="BH")
        if len(xr) > 0:
            ax.scatter(xr[-1], yr[-1], s=30, marker="x")

    ax.scatter([0], [0], s=40, c="gold", label="Sun (fixed)")
    ax.legend(fontsize=8, ncols=2)
    plt.tight_layout()

    out = _prefix_path(base_dir, "heliocentric_plot.png")
    plt.savefig(out, dpi=200)
    plt.close(fig)
    print(f"Saved heliocentric plot: {out}")


def plot_xy_inner3AU_from_logs(base_dir: Path, params: dict, logs: dict, rmax=3.0):
    sun_x = np.array(logs["Sun"]["x"])
    sun_y = np.array(logs["Sun"]["y"])

    def rel_xy(nm):
        x = np.array(logs[nm]["x"])
        y = np.array(logs[nm]["y"])
        n = min(len(x), len(sun_x))
        return x[:n] - sun_x[:n], y[:n] - sun_y[:n]

    fig, ax = plt.subplots(figsize=(8, 8))
    ax.set_aspect("equal")
    ax.grid(True, alpha=0.3)
    ax.set_xlabel("x [AU]")
    ax.set_ylabel("y [AU]")
    ax.set_title(f"Inner Solar System (|x|,|y| ≤ {rmax} AU)")

    # Parameter box
    _add_params_box(ax, params, loc=(0.02, 0.98))

    # 3 AU boundary circle
    theta = np.linspace(0, 2*np.pi, 400)
    ax.plot(
        rmax*np.cos(theta), rmax*np.sin(theta),
        ls="--", lw=0.8, alpha=0.6,
        label=f"{rmax} AU boundary"
    )

    inner_bodies = ["Mercury","Venus","Earth","Mars","Jupiter","Saturn","Uranus","Neptune","Moon","BH"]
    for nm in inner_bodies:
        if nm not in logs or len(logs[nm]["x"]) == 0:
            continue
        xr, yr = rel_xy(nm)
        if len(xr) == 0:
            continue
        ax.plot(xr, yr, lw=1.0, label=nm)
        ax.scatter(xr[-1], yr[-1], s=10)

    ax.scatter([0], [0], s=40, c="gold", label="Sun")
    ax.set_xlim(-rmax, rmax)
    ax.set_ylim(-rmax, rmax)
    ax.legend(fontsize=8, ncols=2)
    plt.tight_layout()

    out = _prefix_path(base_dir, f"heliocentric_inner_{safe_num_for_filename(rmax)}AU.png")
    plt.savefig(out, dpi=200)
    plt.close(fig)
    print(f"Saved inner {rmax} AU heliocentric plot: {out}")

def make_inner_xy_animation(base_dir: Path, logs: dict, inner_names=None, radius_au=3.0, write_mp4=True):
    """
    2D animation (x–y) of inner planets + Moon, clipped to radius_au.
    Saves MP4 via ffmpeg if available; otherwise prints a warning and skips.
    """
    if not write_mp4:
        print("[anim] write_mp4=False → skipping 2D animation")
        return
    
    print("make_inner_xy_animation 1")

    sun_x = np.array(logs["Sun"]["x"])
    sun_y = np.array(logs["Sun"]["y"])
    t = np.array(logs["Sun"]["t"])
    if t.size == 0:
        print("[anim2D] No log data; skipping 2D inner animation.")
        return

    print("make_inner_xy_animation 2")

    rel = {}
    for nm in INNER_BODY_NAMES:
        x = np.array(logs[nm]["x"])
        y = np.array(logs[nm]["y"])
        n = min(len(x), len(sun_x))
        xr = x[:n] - sun_x[:n]
        yr = y[:n] - sun_y[:n]
        r = np.sqrt(xr * xr + yr * yr)
        xr = np.where(r <= radius_au, xr, np.nan)
        yr = np.where(r <= radius_au, yr, np.nan)
        rel[nm] = (xr, yr)
    t = t[: len(sun_x)]
    n_frames = len(t)
    if n_frames < 2:
        print("[anim2D] Not enough frames; skipping.")
        return

    print("make_inner_xy_animation 3")

    fig, ax = plt.subplots(figsize=(6, 6))
    ax.set_aspect("equal")
    ax.grid(True, alpha=0.3)
    ax.set_xlim(-radius_au, radius_au)
    ax.set_ylim(-radius_au, radius_au)
    ax.set_xlabel("x [AU]")
    ax.set_ylabel("y [AU]")
    ax.set_title("Inner planets (2D x–y)")
    ax.scatter([0], [0], s=40, c="gold", label="Sun")

    lines = {}
    markers = {}
    for nm in INNER_BODY_NAMES:
        line, = ax.plot([], [], lw=1.0, label=nm)
        marker = ax.scatter([], [], s=15)
        lines[nm] = line
        markers[nm] = marker

    ax.legend(fontsize=8)

    print("make_inner_xy_animation 4")

    def init():
        for nm in INNER_BODY_NAMES:
            lines[nm].set_data([], [])
            markers[nm].set_offsets(np.array([[np.nan, np.nan]]))
        return list(lines.values()) + [m for m in markers.values()]

    def update(frame):
        for nm in INNER_BODY_NAMES:
            xr, yr = rel[nm]
            lines[nm].set_data(xr[: frame + 1], yr[: frame + 1])
            if frame < len(xr) and np.isfinite(xr[frame]) and np.isfinite(yr[frame]):
                markers[nm].set_offsets(np.array([[xr[frame], yr[frame]]]))
            else:
                markers[nm].set_offsets(np.array([[np.nan, np.nan]]))
        return list(lines.values()) + [m for m in markers.values()]

    ani = animation.FuncAnimation(
        fig,
        update,
        frames=n_frames,
        init_func=init,
        blit=True,
        interval=40,
    )

    out_mp4 = _prefix_path(base_dir, "inner_xy_animation.mp4")
    try:
        Writer = animation.writers["ffmpeg"]
        writer = Writer(fps=25, metadata={"artist": "rebound"}, bitrate=1800)
        ani.save(out_mp4, writer=writer)
        print(f"Saved 2D inner-system animation: {out_mp4}")
    except Exception as e:
        print(f"[anim2D] Could not save MP4 via ffmpeg ({e}); skipping 2D animation.")
    plt.close(fig)


def make_inner_3d_animation(base_dir: Path, logs: dict, inner_names=None, radius_au=3.0, write_mp4=True):
    """
    3D animation of inner planets + Moon, clipped to radius_au in 3D.
    Saves MP4 via ffmpeg if available; otherwise prints a warning and skips.
    """
    if not write_mp4:
        print("[anim] write_mp4=False → skipping 3D animation")
        return

    sun_x = np.array(logs["Sun"]["x"])
    sun_y = np.array(logs["Sun"]["y"])
    sun_z = np.array(logs["Sun"]["z"])
    t = np.array(logs["Sun"]["t"])
    if t.size == 0:
        print("[anim3D] No log data; skipping 3D inner animation.")
        return

    rel = {}
    for nm in INNER_BODY_NAMES:
        x = np.array(logs[nm]["x"])
        y = np.array(logs[nm]["y"])
        z = np.array(logs[nm]["z"])
        n = min(len(x), len(sun_x))
        xr = x[:n] - sun_x[:n]
        yr = y[:n] - sun_y[:n]
        zr = z[:n] - sun_z[:n]
        r = np.sqrt(xr * xr + yr * yr + zr * zr)
        xr = np.where(r <= radius_au, xr, np.nan)
        yr = np.where(r <= radius_au, yr, np.nan)
        zr = np.where(r <= radius_au, zr, np.nan)
        rel[nm] = (xr, yr, zr)

    t = t[: len(sun_x)]
    n_frames = len(t)
    if n_frames < 2:
        print("[anim3D] Not enough frames; skipping.")
        return

    fig = plt.figure(figsize=(7, 7))
    ax = fig.add_subplot(111, projection="3d")
    ax.set_xlim(-radius_au, radius_au)
    ax.set_ylim(-radius_au, radius_au)
    ax.set_zlim(-radius_au, radius_au)
    ax.set_xlabel("x [AU]")
    ax.set_ylabel("y [AU]")
    ax.set_zlabel("z [AU]")
    ax.set_title("Inner planets (3D)")
    ax.scatter([0], [0], [0], s=40, c="gold", label="Sun")

    lines = {}
    markers = {}
    for nm in INNER_BODY_NAMES:
        line, = ax.plot([], [], [], lw=1.0, label=nm)
        marker = ax.scatter([], [], [], s=15)
        lines[nm] = line
        markers[nm] = marker

    # We can't use normal legend easily with scatter+lines here; skip or minimal.
    # ax.legend(fontsize=8)

    def init():
        for nm in INNER_BODY_NAMES:
            lines[nm].set_data([], [])
            lines[nm].set_3d_properties([])
        return list(lines.values()) + [m for m in markers.values()]

    def update(frame):
        for nm in INNER_BODY_NAMES:
            xr, yr, zr = rel[nm]
            lines[nm].set_data(xr[: frame + 1], yr[: frame + 1])
            lines[nm].set_3d_properties(zr[: frame + 1])
            if frame < len(xr) and np.isfinite(xr[frame]) and np.isfinite(yr[frame]) and np.isfinite(zr[frame]):
                markers[nm]._offsets3d = ([xr[frame]], [yr[frame]], [zr[frame]])
            else:
                markers[nm]._offsets3d = ([np.nan], [np.nan], [np.nan])
        return list(lines.values()) + [m for m in markers.values()]

    ani = animation.FuncAnimation(
        fig,
        update,
        frames=n_frames,
        init_func=init,
        blit=False,
        interval=40,
    )

    out_mp4 = _prefix_path(base_dir, "inner_3d_animation.mp4")
    try:
        Writer = animation.writers["ffmpeg"]
        writer = Writer(fps=25, metadata={"artist": "rebound"}, bitrate=1800)
        ani.save(out_mp4, writer=writer)
        print(f"Saved 3D inner-system animation: {out_mp4}")
    except Exception as e:
        print(f"[anim3D] Could not save MP4 via ffmpeg ({e}); skipping 3D animation.")
    plt.close(fig)


def make_belt_figures(base_dir: Path, params: dict, df_b: pd.DataFrame, df_a: pd.DataFrame):
    """Create belt figures. If n_belt==0 (empty dfs), skip gracefully."""
    if df_b is None or df_a is None or len(df_b) == 0 or len(df_a) == 0:
        print("[belt] n_belt=0 (or empty belt snapshots). Skipping belt figures.")
        return

    df = pd.concat([df_b.reset_index(drop=True), df_a.reset_index(drop=True)], axis=1)
    rp_token = safe_num_for_filename(params["bh_rp_au"])
    m_token  = safe_num_for_filename(params["bh_mass_msun"])

    a_b = df["a_before"].to_numpy(); e_b = df["e_before"].to_numpy()
    a_a = df["a_after"].to_numpy();  e_a = df["e_after"].to_numpy()

    def pad(lo, hi, frac=0.05):
        span = hi - lo
        return lo - frac*span, hi + frac*span

    a_all = np.concatenate([a_b, a_a])
    e_all = np.concatenate([e_b, e_a])

    axmin, axmax = pad(np.nanmin(a_all), np.nanmax(a_all))
    eymin, eymax = pad(max(0.0, np.nanmin(e_all)), np.nanmax(e_all))

    fig1, ax1 = plt.subplots(figsize=(7, 5))
    ax1.scatter(a_b, e_b, s=6, alpha=0.35, label="before")
    ax1.scatter(a_a, e_a, s=6, alpha=0.35, label="after")
    ax1.set_xlabel("a [AU]")
    ax1.set_ylabel("e")
    ax1.set_xlim(axmin, axmax)
    ax1.set_ylim(eymin, eymax)
    ax1.grid(True, alpha=0.3)
    ax1.set_title("Asteroid belt: e vs a (before/after)")
    ax1.legend(markerscale=2, frameon=False)
    out1 = _prefix_path(base_dir, f"belt_run__{rp_token}__{m_token}_ae_scatter.png")
    plt.tight_layout()
    fig1.savefig(out1, dpi=160)
    plt.close(fig1)

    q_b = df["a_before"] * (1.0 - df["e_before"])
    q_a = df["a_after"] * (1.0 - df["e_after"])
    da = df["a_after"] - df["a_before"]
    dq = q_a - q_b

    fig2, axes = plt.subplots(2, 1, figsize=(7, 6), sharex=False)
    ax = axes[0]
    ax.hist(da[~np.isnan(da)], bins=60)
    ax.set_xlabel("Δa [AU]")
    ax.set_ylabel("Count")
    ax.set_title("Semimajor axis change (Δa)")
    ax.grid(True, alpha=0.3)

    ax = axes[1]
    ax.hist(dq[~np.isnan(dq)], bins=60)
    ax.set_xlabel("Δq [AU]")
    ax.set_ylabel("Count")
    ax.set_title("Perihelion change (Δq)")
    ax.grid(True, alpha=0.3)

    plt.tight_layout()
    out2 = _prefix_path(base_dir, f"belt_run__{rp_token}__{m_token}_delta_hist.png")
    fig2.savefig(out2, dpi=160)
    plt.close(fig2)

    print(f"Saved {out1.name}, {out2.name}")


# ===========================================================
# Post-processing: belt deltas, planet deltas, sizes, hazard
# ===========================================================
def _write_belt_run_deltas(base_dir: Path, df_b: pd.DataFrame, df_a: pd.DataFrame):
    """Save per-particle before/after + deltas to CSV."""
    q_b = df_b["a_before"] * (1.0 - df_b["e_before"])
    q_a = df_a["a_after"] * (1.0 - df_a["e_after"])
    deltas = pd.DataFrame(
        {
            "a_before": df_b["a_before"],
            "e_before": df_b["e_before"],
            "q_before": q_b,
            "a_after": df_a["a_after"],
            "e_after": df_a["e_after"],
            "q_after": q_a,
        }
    )
    deltas["da"] = deltas["a_after"] - deltas["a_before"]
    deltas["de"] = deltas["e_after"] - deltas["e_before"]
    deltas["dq"] = deltas["q_after"] - deltas["q_before"]
    out = _prefix_path(base_dir, "belt_run_deltas.csv")
    deltas.to_csv(out, index=False)
    return deltas


def _write_planets_run_deltas(base_dir: Path,
                              df_before: pd.DataFrame,
                              df_after: pd.DataFrame,
                              planet_names=PLANET_NAMES_FOR_DELTAS):
    """
    Write the SAME planets_run_deltas.csv format as the older script:

      body,a_before,e_before,q_before,a_after,e_after,q_after,da,de,dq

    Works even if df_before/df_after columns are named with suffixes like:
      a_before_AU, e_before, q_before_AU, etc.
    """

    def _col(df, base):
        """
        Return a Series for a column that might be named:
          base
          base + "_AU"
          base + "_km" (etc)
        We accept any column that starts with base exactly.
        """
        if base in df.columns:
            return df[base]
        # fallback: pick first col that starts with base
        matches = [c for c in df.columns if c.startswith(base)]
        if matches:
            return df[matches[0]]
        raise KeyError(f"Could not find column '{base}' (or '{base}*') in df columns: {list(df.columns)}")

    # Index by body for easy lookup
    df_b = df_before.set_index("body")
    df_a = df_after.set_index("body")

    rows = []
    for b in planet_names:
        if b not in df_b.index or b not in df_a.index:
            continue

        a_b = float(_col(df_b.loc[[b]].reset_index(), "a").iloc[0])
        e_b = float(_col(df_b.loc[[b]].reset_index(), "e").iloc[0])
        q_b = float(_col(df_b.loc[[b]].reset_index(), "q").iloc[0])

        a_a = float(_col(df_a.loc[[b]].reset_index(), "a").iloc[0])
        e_a = float(_col(df_a.loc[[b]].reset_index(), "e").iloc[0])
        q_a = float(_col(df_a.loc[[b]].reset_index(), "q").iloc[0])

        rows.append({
            "body": b,
            "a_before": a_b,
            "e_before": e_b,
            "q_before": q_b,
            "a_after":  a_a,
            "e_after":  e_a,
            "q_after":  q_a,
            "da": a_a - a_b,
            "de": e_a - e_b,
            "dq": q_a - q_b,
        })

    df_out = pd.DataFrame(rows)
    out = _prefix_path(base_dir, "planets_run_deltas.csv")
    df_out.to_csv(out, index=False)
    print(f"Wrote planets deltas to {out}")
    return df_out

def _sample_sizes_km(n, slope, dmin, dmax, seed=1234):
    """
    Sample diameters from a power-law N(>D) ~ D^{-slope} (cumulative slope).
    Inverse-CDF method on cumulative distribution.
    Handles slope ~= 1 separately.
    """
    rng = np.random.default_rng(seed)
    u = rng.random(n)
    b = float(slope)
    dmin = float(dmin)
    dmax = float(dmax)
    if b == 1.0:
        # Log-uniform
        return dmin * (dmax / dmin) ** u
    # For b != 1: F(D) = (D^{1-b} - dmin^{1-b}) / (dmax^{1-b} - dmin^{1-b})
    p = u
    A = dmin ** (1.0 - b)
    B = dmax ** (1.0 - b)
    return (A + p * (B - A)) ** (1.0 / (1.0 - b))


def _crossers_mask(q_after_series, q_thresh_au=1.0):
    return (q_after_series.values < q_thresh_au) & np.isfinite(q_after_series.values)


def _write_crossers_size_histograms(
    base_dir: Path,
    sizes_all_km,
    crossers_mask,
    bins=30,
):
    """Write histogram CSV + PNG + SVG for crossers (using all sampled sizes)."""
    cross_sizes = sizes_all_km[crossers_mask]
    if cross_sizes.size == 0:
        # Write empty CSV, skip plots
        csv_path = _prefix_path(base_dir, "belt_crossers_size_histogram.csv")
        pd.DataFrame(
            {"bin_left_km": [], "bin_right_km": [], "count": []}
        ).to_csv(csv_path, index=False)
        return

    # Use log-spaced bins across the range present in crossers
    lo = cross_sizes.min()
    hi = cross_sizes.max()
    if lo <= 0:
        lo = np.nextafter(0, 1)  # guard
    edges = np.logspace(np.log10(lo), np.log10(hi), bins + 1)
    counts, edges = np.histogram(cross_sizes, bins=edges)

    hist_df = pd.DataFrame(
        {
            "bin_left_km": edges[:-1],
            "bin_right_km": edges[1:],
            "count": counts,
        }
    )
    csv_path = _prefix_path(base_dir, "belt_crossers_size_histogram.csv")
    hist_df.to_csv(csv_path, index=False)

    # Plot (PNG + SVG)
    fig = plt.figure(figsize=(7, 5))
    ax = fig.add_subplot(111)
    ax.hist(cross_sizes, bins=edges)
    ax.set_xscale("log")
    ax.set_xlabel("Diameter [km]")
    ax.set_ylabel("Count")
    ax.set_title("Earth-crossing tracers: size distribution")
    ax.grid(True, which="both", alpha=0.3)
    plt.tight_layout()
    png_path = _prefix_path(base_dir, "belt_crossers_size_histogram.png")
    svg_path = _prefix_path(base_dir, "belt_crossers_size_histogram.svg")
    fig.savefig(png_path, dpi=160)
    fig.savefig(svg_path)
    plt.close(fig)


def _write_hazard_summary(
    base_dir: Path,
    params: dict,
    deltas_df: pd.DataFrame,
    sizes_all_km: np.ndarray,
    crossers_mask: np.ndarray,
):
    """Write hazard summary text with scaling to a real belt."""
    size_thresh = 1.0  # km
    n_total = len(sizes_all_km)
    ge1_all = np.sum(sizes_all_km >= size_thresh)
    q_after = deltas_df["q_after"].values
    ge1_cross = np.sum((sizes_all_km >= size_thresh) & (q_after < 1.0))

    # Real-belt scaling
    real_ge1km = int(params.get("real_ge1km_count", 1_200_000))
    # Fraction among simulated tracers:
    f_cross = (ge1_cross / ge1_all) if ge1_all > 0 else 0.0
    est_real_crossers = int(round(f_cross * real_ge1km))

    # Hazard model (simple Poisson with constant per-object annual probability)
    p = float(params.get("per_object_impact_prob_yr", 1.0e-8))
    R_real = est_real_crossers * p  # expected aggregate impacts per year
    mean_wait = (1.0 / R_real) if R_real > 0 else float("inf")
    T10 = 10.0
    P_le_10 = (1.0 - math.exp(-R_real * T10)) if R_real > 0 else 0.0

    # Write text file
    out = _prefix_path(base_dir, "hazard_summary.txt")
    lines = []
    lines.append("=== Belt Impact Hazard Summary ===")
    lines.append(f"Folder: {base_dir}")
    lines.append(f"Tracers analyzed: {n_total}")
    lines.append(f"Assigned size slope: {params.get('size_slope', 2.5)}")
    lines.append(f"Seed: {params.get('size_seed', 1234)}")
    lines.append("")
    lines.append("--- Sample stats ---")
    lines.append(f"≥1 km tracers                : {ge1_all}")
    lines.append(f"≥1 km Earth-crossers (q<1 AU): {ge1_cross}")
    lines.append(f"Fraction f_cross             : {f_cross:.6f}")
    lines.append("")
    lines.append("--- Real-belt scaling ---")
    lines.append(f"Assumed real ≥1 km count     : {real_ge1km:,}")
    lines.append(f"Estimated real ≥1 km crossers: {est_real_crossers:,}")
    lines.append("")
    lines.append("--- Hazard metrics ---")
    lines.append(f"Per-object impact prob       : {p:.1e} yr^-1")
    lines.append(f"Aggregate annual rate R_real : {R_real:.3e} yr^-1")
    lines.append(
        "Mean waiting time            : "
        + ("∞" if not math.isfinite(mean_wait) else f"{int(round(mean_wait)):,} years")
    )
    lines.append(f"P(≤10 years)             : {100.0 * P_le_10:0.4f} %")
    lines.append("")
    lines.append("(Assumes simple Poisson process with constant R_real.)")
    with open(out, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


# ===========================================================
# One-run wrapper
# ===========================================================
def run_one(params, sub_dir: Path, params_src_path: Path):
    sub_dir.mkdir(parents=True, exist_ok=True)
    # Copy params into subfolder with prefix
    if params_src_path and params_src_path.exists():
        shutil.copy2(params_src_path, _prefix_path(sub_dir, "input.yaml"))
    print(f"Run directory: {sub_dir}")
    print(f"Copied params file to: {_prefix_path(sub_dir, 'input.yaml')}")

    sim, belt, mu_sun_bh = build_sim(params)
    print(f"Simulation built. BH mass = {params['bh_mass_msun']:.3f} Msun.")

    sim_baseline = None
    baseline_mass = params.get("compare_baseline_bh_mass_msun", None)
    if baseline_mass is not None:
        baseline_params = dict(params)
        baseline_params["bh_mass_msun"] = float(baseline_mass)
        baseline_params["n_belt"] = 0
        baseline_params["archive_enable"] = False
        baseline_params["write_npz"] = False
        sim_baseline, _belt_unused, _mu_unused = build_sim(baseline_params)
        print(f"Baseline simulation built. BH mass = {float(baseline_mass):.3f} Msun.")

    n_belt = int(params.get("n_belt", 0))
    has_belt = n_belt > 0

    # Optional SimulationArchive
    archive_enable = bool(params.get("archive_enable", False))
    if archive_enable:
        arch_interval = float(
            params.get(
                "archive_interval_days",
                params.get("dump_interval_days", 10.0),
            )
        )
        arch_name = params.get("archive_filename", "archive.bin")
        archive_path = _prefix_path(sub_dir, arch_name)

        if hasattr(sim, "save_to_file"):
            try:
                sim.save_to_file(
                    str(archive_path),
                    interval=arch_interval,
                    delete_file=True,
                )
                print(
                    f"[archive] SimulationArchive to {archive_path} "
                    f"every {arch_interval} days"
                )
            except Exception as e:
                print(
                    f"[archive] could not enable SimulationArchive via save_to_file: {e}"
                )
        else:
            print(
                "[archive] sim.save_to_file() not available in this REBOUND build; "
                "archive disabled."
            )

    # Belt BEFORE
    if has_belt:
        df_b = snapshot_belt(sim, belt["start"], belt["count"], "before", mu=mu_sun_bh)
        df_b.to_csv(_prefix_path(sub_dir, "belt_run_before.csv"), index=False)
        print("Wrote belt_run_before.csv")
    else:
        df_b = None

    # Planets BEFORE
    # Use Sun-only gravitational parameter for heliocentric osculating elements
    mu_planets = sim.G * MSUN
    planets_before = snapshot_planets_heliocentric(
        sim,
        mu_planets,
        body_names=PLANET_NAMES_FOR_DELTAS,
    )

    # Logs for Excel / plots / animations
    logs = init_logs(BODY_NAMES)

    epoch_dt_utc = parse_epoch_utc(params["epoch"])  # timezone-aware UTC datetime

    write_npz = bool(params.get("write_npz", True))
    bh_radec = integrate_and_dump(
        sim,
        duration_days=params["duration_days"],
        dump_interval_days=params["dump_interval_days"],
        output_interval_days=params["output_interval_days"],
        output_dir=sub_dir,
        body_names=BODY_NAMES,
        logs=logs,
        epoch_dt_utc=parse_epoch_utc(params["epoch"]),  # or however you already pass it
        write_npz=write_npz,
        sim_baseline=sim_baseline,
    )

    # Belt AFTER
    if has_belt:
        df_a = snapshot_belt(sim, belt["start"], belt["count"], "after", mu=mu_sun_bh)
        df_a.to_csv(_prefix_path(sub_dir, "belt_run_after.csv"), index=False)
        print("Wrote belt_run_after.csv")
    else:
        df_a = None

    # Planets AFTER
    planets_after = snapshot_planets_heliocentric(
        sim,
        mu_planets,
        body_names=PLANET_NAMES_FOR_DELTAS,
    )

    # Planets deltas
    _write_planets_run_deltas(sub_dir, planets_before, planets_after)

    # Figures (ae scatter + delta hist for belt)
    if has_belt:
        make_belt_figures(sub_dir, params, df_b, df_a)
    else:
        print("[belt] n_belt=0 → skipping belt figures")

    # Excel + heliocentric plots
    write_excel(sub_dir, params, logs, BODY_NAMES)

    # BH RA/Dec Excel
    write_bh_radec_excel(sub_dir, params, bh_radec)

    # CSV(s) for estimate_BH_parameters_uncertainties_fast.py
    if bool(params.get("write_uncertainty_csv", True)):
        write_uncertainty_csv(
            sub_dir,
            params,
            logs,
            body_name=str(params.get("uncertainty_body", "Earth")),
        )

    plot_xy_heliocentric_from_logs(sub_dir, params, logs, BODY_NAMES)
    plot_xy_inner3AU_from_logs(sub_dir, params, logs)

    # Deltas, sizes, hazard for belt
    if has_belt:
        deltas_df = _write_belt_run_deltas(sub_dir, df_b, df_a)

        slope = float(params.get("size_slope", 2.5))
        dmin = float(params.get("size_min_km", 0.1))
        dmax = float(params.get("size_max_km", 50.0))
        seed = int(params.get("size_seed", 1234))

        sizes_km = _sample_sizes_km(len(deltas_df), slope, dmin, dmax, seed=seed)
        cross_mask = _crossers_mask(deltas_df["q_after"], q_thresh_au=1.0)

        _write_crossers_size_histograms(sub_dir, sizes_km, cross_mask, bins=30)
        _write_hazard_summary(sub_dir, params, deltas_df, sizes_km, cross_mask)
    else:
        print("[belt] n_belt=0 → skipping belt deltas & hazard products")

    # Inner-system animations
    write_mp4 = bool(params.get("write_mp4", True))
    make_inner_xy_animation(sub_dir, logs, write_mp4=write_mp4)
    make_inner_3d_animation(sub_dir, logs, write_mp4=write_mp4)

    print("Done.")


def _run_one_worker(params, sub_dir: Path, params_src_path: Path, sub_name: str):
    """
    Picklable top-level wrapper around run_one() for use with
    ProcessPoolExecutor. Each combo in a sweep is fully independent (its own
    Simulation, its own sub_dir), so this isolates one combo's failure from
    the rest of the batch instead of letting one bad case kill the whole
    sweep, and reports success/failure back to the main process.
    """
    print(f"[worker] Starting: {sub_name}")
    try:
        run_one(params, sub_dir, params_src_path)
        print(f"[worker] Finished: {sub_name}")
        return sub_name, True, None
    except Exception:
        err = traceback.format_exc()
        try:
            print(f"[worker] FAILED: {sub_name}\n{err}")
        except UnicodeEncodeError:
            print(f"[worker] FAILED: {sub_name} (see returned traceback; console could not display it)")
        return sub_name, False, err


def _worker_init():
    """
    ProcessPoolExecutor initializer, run once per spawned worker process.

    A freshly spawned worker on Windows does not inherit the main process's
    console/UTF-8 stdout configuration and can default to a legacy codepage
    (e.g. cp1252). Several print() calls in this module use non-ASCII
    characters (e.g. "->" as "→", "Omega"/"omega" as Greek letters), which
    would otherwise crash the worker with UnicodeEncodeError the first time
    one is printed.
    """
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass


# ===========================================================
# Main (with resume support)
# ===========================================================
def main():
    # Some print() calls below use non-ASCII characters (e.g. "→"). Console
    # stdout is usually UTF-8-capable, but redirecting to a file/pipe can fall
    # back to the system codepage (cp1252 on this machine) and crash on them.
    _worker_init()

    ap = argparse.ArgumentParser(
        description="Solar System + BH + Belt (params from YAML)"
    )
    ap.add_argument("params", help="YAML parameter file (e.g., input.yaml)")
    ap.add_argument(
        "--resume-dir",
        type=str,
        default=None,
        help=(
            "Existing simulations/<STAMP> folder to resume in. "
            "Deletes the last completed subfolder and resumes from there."
        ),
    )
    ap.add_argument(
        "--workers",
        type=int,
        default=max(1, (os.cpu_count() or 2) - 1),
        help=(
            "Number of sweep combinations to run in parallel (separate processes). "
            "Each combo is fully independent, so this only matters when the YAML "
            "defines ranges (e.g. bh_rp_au_range). "
            "1 = original sequential behavior. Default: cpu_count - 1."
        ),
    )
    ap.add_argument(
        "--max-worker-mem-gb",
        type=float,
        default=2.5,
        help=(
            "Estimated peak RAM (GB) used by a single worker process (ephemeris + "
            "IAS15 state + belt tracers). Used to cap --workers so total memory demand "
            "stays under available system RAM. Default: 2.5 GB/worker."
        ),
    )
    args = ap.parse_args()

    if args.workers > 1:
        if psutil is not None:
            available_gb = psutil.virtual_memory().available / (1024 ** 3)
            # Leave headroom for the OS/other apps; only use 80% of what's free.
            mem_cap = max(1, int((available_gb * 0.8) // args.max_worker_mem_gb))
        else:
            # No psutil: fall back to a conservative cap since we can't see RAM.
            mem_cap = max(1, (os.cpu_count() or 2) // 2)
            print(
                "[safety] psutil not installed -- cannot check available RAM; "
                f"capping --workers at {mem_cap} as a precaution."
            )
        if args.workers > mem_cap:
            print(
                f"[safety] Requested --workers {args.workers} exceeds the memory-safe cap "
                f"of {mem_cap} (est. {args.max_worker_mem_gb} GB/worker). "
                f"Reducing to {mem_cap} to avoid exhausting system memory. "
                f"Override with --max-worker-mem-gb if your combos use less RAM than that."
            )
            args.workers = mem_cap

    with open(args.params, "r", encoding="utf-8") as f:
        params_master = yaml.safe_load(f)

    # Determine sweep vectors (backward compatible with single OR range values)
    def _get_vals(master, val_key, range_key):
        if range_key in master:
            return _parse_range(master[range_key])
        raw = master.get(val_key)
        if isinstance(raw, (list, tuple)) or (isinstance(raw, str) and "," in raw):
            return _parse_range(raw)
        return np.array([float(raw)])

    rp_vals = _get_vals(params_master, "bh_rp_au", "bh_rp_au_range")
    vinf_vals = _get_vals(params_master, "bh_vinf_kms", "bh_vinf_kms_range")
    inc_vals = _get_vals(params_master, "bh_inc_deg", "bh_inc_deg_range")
    toff_vals = _get_vals(
        params_master,
        "bh_tperi_offset_days",
        "bh_tperi_offset_days_range",
    )
    Omega_vals = _get_vals(params_master, "bh_Omega_deg", "bh_Omega_deg_range")
    omega_vals = _get_vals(params_master, "bh_omega_deg", "bh_omega_deg_range")

    # Root timestamp and directory
    if args.resume_dir is not None:
        # Use existing folder to resume from
        root_dir = Path(args.resume_dir)
        if not root_dir.exists():
            # Allow passing just the stamp name, e.g., "20251127_201847"
            alt = Path("simulations") / args.resume_dir
            if alt.exists():
                root_dir = alt
            else:
                raise FileNotFoundError(f"Resume directory not found: {args.resume_dir}")
        stamp = root_dir.name
        print(f"[resume] Using existing root directory: {root_dir}")
    else:
        # Fresh run
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        root_dir = Path("simulations") / stamp
        root_dir.mkdir(parents=True, exist_ok=True)
        print(f"[new] Root directory: {root_dir}")

    # Build ordered list of all parameter combinations and their subfolder names
    combos = []
    for rp, vinf, inc, toff, Om, om in product(
        rp_vals, vinf_vals, inc_vals, toff_vals, Omega_vals, omega_vals
    ):
        rp_token = safe_num_for_filename(rp)
        vinf_token = safe_num_for_filename(vinf)
        inc_token = safe_num_for_filename(inc)
        toff_token = safe_num_for_filename(toff)
        Om_token = safe_num_for_filename(Om)
        om_token = safe_num_for_filename(om)

        sub_name = (
            f"{stamp}__rp{rp_token}"
            f"__vinf{vinf_token}"
            f"__inc{inc_token}"
            f"__toff{toff_token}"
            f"__Om{Om_token}"
            f"__om{om_token}"
        )

        combos.append((rp, vinf, inc, toff, Om, om, sub_name))

    # Decide where to start in the combos list
    start_index = 0
    if args.resume_dir is not None:
        existing_dirs = {p.name for p in root_dir.iterdir() if p.is_dir()}
        last_existing_idx = -1
        for i, (_, _, _, _, _, _, sub_name) in enumerate(combos):
            if sub_name in existing_dirs:
                last_existing_idx = i
            else:
                break

        if last_existing_idx >= 0:
            last_sub_name = combos[last_existing_idx][6]
            last_sub_path = root_dir / last_sub_name
            print(f"[resume] Last completed run appears to be: {last_sub_name}")
            print(
                f"[resume] Deleting this folder so its case will be re-run..."
            )
            shutil.rmtree(last_sub_path, ignore_errors=False)
            start_index = last_existing_idx
        else:
            print(
                "[resume] No existing subruns detected that match this parameter grid; "
                "starting from the first combination."
            )
            start_index = 0

    # Build the list of remaining jobs (each combo is fully independent: its
    # own params dict, its own sub_dir, no shared state between combos).
    jobs = []
    for rp, vinf, inc, toff, Om, om, sub_name in combos[start_index:]:
        params = dict(params_master)  # shallow copy is fine
        params["bh_rp_au"] = float(rp)
        params["bh_vinf_kms"] = float(vinf)
        params["bh_inc_deg"] = float(inc)
        params["bh_tperi_offset_days"] = float(toff)
        params["bh_Omega_deg"] = float(Om)
        params["bh_omega_deg"] = float(om)
        jobs.append((params, root_dir / sub_name, sub_name))

    total = len(jobs)
    params_src_path = Path(args.params)

    if args.workers <= 1 or total <= 1:
        # Original sequential behavior: a failure in one combo stops the run.
        for idx, (params, sub_dir, sub_name) in enumerate(jobs, start=1):
            print(f"[{idx}/{total}] Running: {sub_name}")
            run_one(params, sub_dir, params_src_path)
        return

    # Parallel sweep: one process per combo, up to --workers at a time.
    # Unlike the sequential path, a failed combo does not abort the others --
    # all combos run to completion and failures are reported at the end.
    print(f"[parallel] Running {total} combo(s) across up to {args.workers} worker process(es)...")
    failures = []
    with ProcessPoolExecutor(max_workers=args.workers, initializer=_worker_init) as executor:
        futures = {
            executor.submit(_run_one_worker, params, sub_dir, params_src_path, sub_name): sub_name
            for params, sub_dir, sub_name in jobs
        }
        try:
            for idx, future in enumerate(as_completed(futures), start=1):
                sub_name, ok, err = future.result()
                print(f"[{idx}/{total}] {'OK' if ok else 'FAILED'}: {sub_name}")
                if not ok:
                    failures.append((sub_name, err))
        except KeyboardInterrupt:
            print("\n[parallel] Interrupted; cancelling remaining combos...")
            executor.shutdown(cancel_futures=True)
            raise

    if failures:
        print(f"\n[parallel] {len(failures)}/{total} combo(s) failed:")
        for sub_name, _ in failures:
            print(f"  - {sub_name}")
        raise RuntimeError(
            f"{len(failures)} of {total} sweep combination(s) failed. "
            f"See per-combo tracebacks above."
        )


if __name__ == "__main__":
    _script_start = time.time()
    try:
        main()
    finally:
        _elapsed = time.time() - _script_start
        print(f"[timing] Total execution time: {timedelta(seconds=round(_elapsed))} ({_elapsed:.1f} s)")
